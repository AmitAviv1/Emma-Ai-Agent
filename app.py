"""
app.py — Streamlit UI for the Invoice Automation Agent

Run:
    streamlit run app.py

Install:
    pip install streamlit
"""

import os
import sys
import hmac
import tempfile
import streamlit as st
import streamlit.components.v1 as components
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

# ── add nodes/ to path so imports work from project root ──────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "nodes"))

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Invoice Agent",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    /* Desktop sidebar */
    [data-testid="stSidebar"] { min-width: 220px; max-width: 220px; }
    .block-container { padding-top: 2rem; }
    div[data-testid="metric-container"] { background: #f8f8f8; border-radius: 8px; padding: 1rem; }
    .product-row { padding: 10px 0; border-bottom: 0.5px solid #eee; }
    .badge-approved { background:#eaf3de; color:#27500a; padding:2px 10px;
                      border-radius:12px; font-size:12px; font-weight:500; }
    .badge-pending  { background:#faeeda; color:#633806; padding:2px 10px;
                      border-radius:12px; font-size:12px; font-weight:500; }
    .badge-skipped  { background:#f1efe8; color:#5f5e5a; padding:2px 10px;
                      border-radius:12px; font-size:12px; font-weight:500; }

    /* Mobile */
    @media (max-width: 768px) {
        [data-testid="stSidebar"] { min-width: unset; max-width: unset; }
        .block-container { padding: 1rem 0.75rem 2rem; }

        /* Stack all columns vertically */
        [data-testid="column"] {
            width: 100% !important;
            flex: 1 1 100% !important;
            min-width: 100% !important;
        }

        /* Bigger tap targets for buttons */
        .stButton > button {
            min-height: 48px;
            font-size: 16px;
            width: 100%;
        }

        /* Bigger inputs */
        .stTextInput > div > div > input,
        .stSelectbox > div > div {
            font-size: 16px;
            min-height: 44px;
        }

        /* File uploader easier to tap */
        [data-testid="stFileUploader"] {
            padding: 1rem;
        }

        /* Metrics stack nicely */
        div[data-testid="metric-container"] {
            padding: 0.6rem;
        }

        /* Images full width */
        [data-testid="stImage"] img {
            width: 100% !important;
        }

        /* Space for fixed bottom nav */
        .block-container { padding-bottom: 80px !important; }

        /* Hide table header row on very small screens */
        .product-row { font-size: 14px; }
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  PASSWORD GATE
# ─────────────────────────────────────────────
#  Set APP_PASSWORD in .env (local) or Streamlit "Secrets" (cloud) to require
#  a password. If it's not set, the app stays open (convenient for local dev).

def _get_app_password() -> str:
    try:
        if "APP_PASSWORD" in st.secrets:
            return str(st.secrets["APP_PASSWORD"])
    except Exception:
        pass
    return os.getenv("APP_PASSWORD", "")


def _require_login():
    password = _get_app_password()
    if not password:
        return                       # no password configured → app is open
    if st.session_state.get("_auth_ok"):
        return

    def _check():
        entered = st.session_state.get("_pw_input", "")
        st.session_state["_auth_ok"] = hmac.compare_digest(entered, password)
        st.session_state.pop("_pw_input", None)

    st.markdown("## 🔒 Emma — כניסה")
    st.text_input("סיסמה", type="password", key="_pw_input", on_change=_check)
    if st.session_state.get("_auth_ok") is False:
        st.error("סיסמה שגויה, נסה שוב.")
    st.stop()


_require_login()


# ─────────────────────────────────────────────
#  SESSION STATE DEFAULTS
# ─────────────────────────────────────────────

for key, default in {
    "page":            "home",
    "products":        [],
    "vendor_name":     "",
    "invoice_number":  "",
    "invoice_date":    "",
    "approved":        [],
    "pending":         [],
    "decisions":       {},   # barcode → "approve" | "skip", name edits
    "saved":           False,
    "save_result":     "",
    "log_lines":       [],
    "_wolt_loaded":    False,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# Auto-load Wolt catalog once per session if path is configured
if not st.session_state._wolt_loaded:
    try:
        from wolt_catalog import load_catalog, WOLT_CATALOG_PATH
        if WOLT_CATALOG_PATH:
            load_catalog()
        st.session_state._wolt_loaded = True
    except Exception:
        st.session_state._wolt_loaded = True  # don't retry on error


# ─────────────────────────────────────────────
#  SIDEBAR NAVIGATION
# ─────────────────────────────────────────────

def nav_button(label, page, badge=None):
    is_active = st.session_state.page == page
    cols = st.sidebar.columns([4, 1]) if badge else [st.sidebar]
    with cols[0]:
        if st.button(
            label,
            key=f"nav_{page}",
            use_container_width=True,
            type="primary" if is_active else "secondary",
        ):
            st.session_state.page = page
            st.rerun()
    if badge:
        with cols[1]:
            st.markdown(
                f'<div style="background:#e24b4a;color:#fff;border-radius:10px;'
                f'text-align:center;font-size:11px;font-weight:600;padding:2px 6px;'
                f'margin-top:6px">{badge}</div>',
                unsafe_allow_html=True,
            )

with st.sidebar:
    st.markdown("### 🧾 Invoice Agent")
    st.markdown("---")
    nav_button("⌂  Home", "home")
    nav_button("📤  Upload", "upload")
    # Review / Results only appear while an invoice extraction is in progress
    if st.session_state.products or st.session_state.page in ("review", "results"):
        pending_count = len([
            b for b, d in st.session_state.decisions.items()
            if d.get("decision") == "pending"
        ]) or len(st.session_state.pending)
        nav_button(
            "🔍  Review",
            "review",
            badge=pending_count if pending_count > 0 and not st.session_state.saved else None,
        )
        nav_button("📋  Results", "results")
    nav_button("📚  Invoice Catalog", "catalog")
    nav_button("🛒  Catalog", "wolt")
    nav_button("📦  Wolt Export", "wolt_export")
    nav_button("🖼️  Image Processor", "image_processor")
    nav_button("🔎  Product Extractor", "product_extract")
    nav_button("🏢  Suppliers", "suppliers")
    st.markdown("---")
    st.caption("Logs")
    if st.session_state.log_lines:
        st.code("\n".join(st.session_state.log_lines[-20:]), language=None)


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def add_log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    st.session_state.log_lines.append(f"{ts}  {msg}")


def run_extraction(pdf_path: str):
    """Runs the extraction pipeline and populates session state."""
    add_log(f"Starting extraction: {os.path.basename(pdf_path)}")

    # Import here so errors surface clearly in the UI
    try:
        from nodes.extract import extract_invoice_data
    except ImportError:
        try:
            from extract import extract_invoice_data
        except ImportError as e:
            raise ImportError(f"Cannot import extract module: {e}. Make sure extract.py is in nodes/ or the project root.")

    try:
        from nodes.storage import validate_products
    except ImportError:
        try:
            from storage import validate_products
        except ImportError as e:
            raise ImportError(f"Cannot import storage module: {e}.")

    with st.spinner("Extracting invoice data…"):
        result = extract_invoice_data({"file_path": pdf_path})

    products    = result.get("products", [])
    vendor_name = result.get("vendor_name", "Unknown")
    warnings    = result.get("warnings", [])

    add_log(f"Extracted {len(products)} products from {vendor_name}")
    for w in warnings:
        add_log(f"⚠ {w}")

    validation = validate_products(products, vendor_name)

    st.session_state.products       = products
    st.session_state.vendor_name    = vendor_name
    st.session_state.invoice_number = result.get("invoice_number", "")
    st.session_state.invoice_date   = result.get("invoice_date", "")
    st.session_state.approved       = validation["approved"]
    st.session_state.pending        = validation["pending_review"]
    st.session_state.saved          = False
    st.session_state.save_result    = ""

    # Initialise decisions keyed by global product index (barcodes can repeat or be empty)
    decisions = {}
    for i, p in enumerate(validation["approved"]):
        decisions[i] = {"decision": "approve", "name": p["name"], "barcode": p.get("barcode", "")}
    offset = len(validation["approved"])
    pending_with_idx = []
    for i, p in enumerate(validation["pending_review"]):
        idx = offset + i
        decisions[idx] = {"decision": "pending", "name": p["name"], "barcode": p.get("barcode", "")}
        pending_with_idx.append({**p, "_idx": idx})   # attach index to each pending item
    st.session_state.decisions = decisions
    st.session_state.pending   = pending_with_idx     # pending now carries its own index

    add_log(f"Auto-approved {len(validation['approved'])}, pending review: {len(validation['pending_review'])}")


# ─────────────────────────────────────────────
#  INVOICE QUEUE (upload several, process one at a time)
# ─────────────────────────────────────────────

def _process_queued_invoice(idx: int) -> bool:
    """Extract the queued invoice at `idx` into session state. Returns True on
    success. Navigation is left to the caller."""
    queue = st.session_state.get("_invoice_queue", [])
    if idx < 0 or idx >= len(queue):
        return False
    item = queue[idx]
    suffix = f".{item['ext']}"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(item["bytes"])
        tmp_path = tmp.name
    st.session_state._extract_error = False
    try:
        run_extraction(tmp_path)
    except Exception as e:
        st.session_state._extract_error = True
        st.error(f"Extraction failed for {item['name']}: {e}")
        add_log(f"❌ Extraction error ({item['name']}): {e}")
        import traceback
        add_log(traceback.format_exc())
        return False
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
    st.session_state._invoice_queue_idx = idx
    return True


def _goto_after_extraction():
    """Route to review (if items need review) or results."""
    if st.session_state.products:
        st.session_state.page = "review" if st.session_state.pending else "results"
        st.rerun()


def _advance_invoice_queue():
    """Move to the next queued invoice; clears the queue when finished."""
    queue = st.session_state.get("_invoice_queue", [])
    nxt = st.session_state.get("_invoice_queue_idx", 0) + 1
    if nxt >= len(queue):
        st.session_state._invoice_queue = []
        st.session_state._invoice_queue_idx = 0
        st.session_state.page = "upload"
        st.rerun()
        return
    # Advance the pointer first so a failing invoice can be skipped, not retried.
    st.session_state._invoice_queue_idx = nxt
    if _process_queued_invoice(nxt):
        _goto_after_extraction()
    else:
        st.rerun()


def _queue_status_banner():
    """Show 'Invoice X of N' progress (+ a Skip button) when a queue is active."""
    queue = st.session_state.get("_invoice_queue", [])
    if len(queue) <= 1:
        return
    idx = st.session_state.get("_invoice_queue_idx", 0)
    remaining = len(queue) - idx - 1
    cols = st.columns([3, 1])
    cols[0].info(f"📚 Invoice **{idx + 1} of {len(queue)}** — {queue[idx]['name']}"
                 + (f"  ·  {remaining} left in queue" if remaining else "  ·  last one"))
    if remaining and cols[1].button("Skip →", use_container_width=True, key="_queue_skip",
                                     help="Skip this invoice without saving and go to the next"):
        _advance_invoice_queue()


def _post_save_actions():
    """After an invoice is saved: advance the queue (prominent) or finish."""
    queue = st.session_state.get("_invoice_queue", [])
    idx = st.session_state.get("_invoice_queue_idx", 0)
    if len(queue) > 1 and idx + 1 < len(queue):
        if st.button(f"Next invoice ({idx + 2} of {len(queue)}) →",
                     type="primary", use_container_width=True, key="_post_next"):
            _advance_invoice_queue()
    else:
        done = len(queue) > 1
        label = "Finish ✓ — all invoices done" if done else "Process another invoice"
        if st.button(label, use_container_width=True, key="_post_finish"):
            st.session_state._invoice_queue = []
            st.session_state._invoice_queue_idx = 0
            st.session_state.page = "upload"
            st.rerun()


def save_to_sheets():
    """Writes approved products to Google Sheets."""
    from storage import (
        _get_sheet_client, _get_or_create_sheet,
        _write_invoice_lines, _update_products_tab,
        _write_invoice_summary, _vendor_tab_name,
        INVOICE_LINES_HEADERS, PRODUCTS_HEADERS,
    )
    from extract import learn_from_approved

    sheet_id = os.getenv("GOOGLE_SHEET_ID")
    if not sheet_id:
        return "❌ GOOGLE_SHEET_ID not set in .env"

    decisions    = st.session_state.decisions
    vendor_name  = st.session_state.vendor_name
    all_products = st.session_state.products

    # Build final approved list respecting UI decisions + name edits
    approved      = []
    newly_learned = []
    pending_barcodes = {p["barcode"] for p in st.session_state.pending}

    for i, p in enumerate(all_products):
        dec = decisions.get(i, {})
        if dec.get("decision") == "approve":
            edited_name = dec.get("name", p["name"])
            approved.append({**p, "name": edited_name, "status": "approved"})
            if p.get("barcode") in pending_barcodes:
                newly_learned.append({**p, "name": edited_name})

    if not approved:
        return "⚠️ No products approved — nothing to save."

    invoice_meta = {
        "invoice_number": st.session_state.invoice_number,
        "invoice_date":   st.session_state.invoice_date or datetime.now().strftime("%Y-%m-%d"),
        "vendor_name":    vendor_name,
    }

    try:
        client       = _get_sheet_client()
        lines_tab    = _vendor_tab_name(vendor_name, "Invoice Lines")
        products_tab = _vendor_tab_name(vendor_name, "Products")

        ws_lines    = _get_or_create_sheet(client, sheet_id, lines_tab,    INVOICE_LINES_HEADERS)
        ws_products = _get_or_create_sheet(client, sheet_id, products_tab, PRODUCTS_HEADERS)

        _write_invoice_lines(ws_lines, approved, invoice_meta)
        _update_products_tab(ws_products, approved, vendor_name)
        _write_invoice_summary(client, sheet_id, approved, invoice_meta)

        # Auto-learn newly approved unknown products
        if newly_learned:
            learned = learn_from_approved(newly_learned)
            add_log(f"🧠 Learned {learned} new product(s) into catalog")

        add_log(f"✅ Saved {len(approved)} products to Google Sheets")
        return f"✅ Saved {len(approved)} products to Google Sheets."

    except Exception as e:
        add_log(f"❌ Save failed: {e}")
        return f"❌ Save failed: {e}"


# ─────────────────────────────────────────────
#  PAGE: UPLOAD
# ─────────────────────────────────────────────

def page_upload():
    st.title("Upload invoices")

    uploaded_files = st.file_uploader(
        "Drop one or more PDF invoices or images here",
        type=["pdf", "jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded_files:
        st.caption(f"**{len(uploaded_files)}** file(s) selected:")
        for f in uploaded_files:
            st.write(f"• {f.name} — {f.size // 1024} KB")

        n = len(uploaded_files)
        btn_label = "Process invoice" if n == 1 else f"Process {n} invoices (one at a time)"
        if st.button(btn_label, type="primary", use_container_width=True):
            # Snapshot bytes now — the uploader widget resets across reruns.
            st.session_state._invoice_queue = [
                {"name": f.name,
                 "ext": f.name.rsplit(".", 1)[-1].lower(),
                 "bytes": f.getvalue()}
                for f in uploaded_files
            ]
            st.session_state._invoice_queue_idx = 0
            if _process_queued_invoice(0):
                if st.session_state.products:
                    _goto_after_extraction()
                elif not st.session_state.get("_extract_error"):
                    st.warning("No products were extracted. Check the logs in the sidebar.")

    # Show last extraction summary if available
    if st.session_state.products:
        st.markdown("---")
        st.subheader("Last extraction")
        c1, c2, c3, c4 = st.columns(4)
        total_cost = sum(p.get("cost", 0) for p in st.session_state.approved)
        c1.metric("Supplier",  st.session_state.vendor_name)
        c2.metric("Products",  len(st.session_state.products))
        c3.metric("Pending review", len(st.session_state.pending))
        c4.metric("Total (net)", f"₪{total_cost:,.2f}")


# ─────────────────────────────────────────────
#  PAGE: REVIEW
# ─────────────────────────────────────────────

def page_review():
    st.title("Review unknown products")
    _queue_status_banner()

    pending = st.session_state.pending
    if not pending:
        st.markdown("""
        <div style="text-align:center;padding:3rem 1rem">
            <div style="font-size:3.5rem">✅</div>
            <p style="font-size:1.1rem;font-weight:600;color:#444;margin:0.5rem 0">All products recognized</p>
            <p style="color:#999;font-size:0.9rem">Every item was auto-approved from your catalog.</p>
        </div>""", unsafe_allow_html=True)
        if st.button("Go to results →", type="primary", use_container_width=True):
            st.session_state.page = "results"
            st.rerun()
        return

    if st.session_state.saved:
        st.success(st.session_state.save_result)
        _post_save_actions()
        return

    st.caption(
        f"{len(pending)} product(s) not found in catalog. "
        "Approve to save, skip to discard, or edit the name before approving."
    )
    st.markdown("---")

    decisions = st.session_state.decisions

    for p in pending:
        bc  = p.get("barcode", "")
        idx = p["_idx"]   # guaranteed unique — set during extraction
        dec = decisions.get(idx, {"decision": "pending", "name": p["name"], "barcode": bc})

        with st.container():
            col_info, col_name, col_action = st.columns([2, 3, 2])

            with col_info:
                st.markdown("**Barcode**")
                st.code(bc if bc else "—", language=None)
                st.caption(f"Qty: {p['quantity']}  |  Cost: ₪{p['cost']:.2f}")

            with col_name:
                st.markdown("**Product name**")
                edited = st.text_input(
                    "name",
                    value=dec.get("name", p["name"]),
                    key=f"name_{idx}",
                    label_visibility="collapsed",
                )
                decisions[idx] = {**dec, "name": edited}

            with col_action:
                st.markdown("**Decision**")
                current = dec.get("decision", "pending")
                choice = st.radio(
                    "decision",
                    ["approve", "skip"],
                    index=0 if current == "approve" else 1,
                    key=f"dec_{idx}",
                    horizontal=True,
                    label_visibility="collapsed",
                )
                decisions[idx] = {**decisions[idx], "decision": choice}

        st.markdown('<hr style="margin:4px 0;opacity:0.2">', unsafe_allow_html=True)

    st.session_state.decisions = decisions
    st.markdown("---")

    approved_count = sum(1 for d in decisions.values() if d.get("decision") == "approve")
    skipped_count  = sum(1 for d in decisions.values() if d.get("decision") == "skip")
    pending_count  = sum(1 for d in decisions.values() if d.get("decision") == "pending")

    st.caption(f"Approved: {approved_count}  |  Skipped: {skipped_count}  |  Still pending: {pending_count}")

    col_save, col_skip = st.columns([2, 1])
    with col_save:
        if st.button("Save all to Google Sheets", type="primary", use_container_width=True,
                     disabled=pending_count > 0):
            with st.spinner("Saving…"):
                result = save_to_sheets()
            st.session_state.save_result = result
            st.session_state.saved = True
            st.session_state.page = "results"   # show the invoice total + next-invoice action
            st.rerun()
    with col_skip:
        if st.button("Save only auto-approved", use_container_width=True):
            for idx, d in decisions.items():
                if d.get("decision") == "pending":
                    decisions[idx] = {**d, "decision": "skip"}
            st.session_state.decisions = decisions
            with st.spinner("Saving…"):
                result = save_to_sheets()
            st.session_state.save_result = result
            st.session_state.saved = True
            st.session_state.page = "results"   # show the invoice total + next-invoice action
            st.rerun()

    if pending_count > 0:
        st.caption("⬆ Resolve all pending items before saving, or use 'Save only auto-approved'.")


# ─────────────────────────────────────────────
#  PAGE: RESULTS
# ─────────────────────────────────────────────

def page_results():
    st.title("Extraction results")
    _queue_status_banner()

    products = st.session_state.products
    if not products:
        st.markdown("""
        <div style="text-align:center;padding:3rem 1rem">
            <div style="font-size:3.5rem">📋</div>
            <p style="font-size:1.1rem;font-weight:600;color:#444;margin:0.5rem 0">No results yet</p>
            <p style="color:#999;font-size:0.9rem">Upload an invoice to see extracted products here.</p>
        </div>""", unsafe_allow_html=True)
        if st.button("Go to Upload →", type="primary", use_container_width=True):
            st.session_state.page = "upload"
            st.rerun()
        return

    decisions   = st.session_state.decisions
    vendor_name = st.session_state.vendor_name
    total_cost  = sum(p.get("cost", 0) for p in products)

    # Summary metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Supplier",     vendor_name)
    c2.metric("Products",     len(products))
    c3.metric("Invoice total", f"₪{total_cost:,.2f}")
    c4.metric("Saved",        "Yes ✓" if st.session_state.saved else "Not yet")

    inv_num  = st.session_state.get("invoice_number", "")
    inv_date = st.session_state.get("invoice_date", "")
    if inv_num or inv_date:
        st.caption(f"Invoice: **{inv_num}**   Date: **{inv_date}**")

    if st.session_state.save_result:
        if "✅" in st.session_state.save_result:
            st.success(st.session_state.save_result)
        else:
            st.error(st.session_state.save_result)

    st.markdown("---")

    # Filter bar
    col_search, col_filter = st.columns([3, 1])
    with col_search:
        search = st.text_input("Search", placeholder="Product name or barcode…",
                               label_visibility="collapsed")
    with col_filter:
        show = st.selectbox("Show", ["All", "Approved", "Pending", "Skipped"],
                            label_visibility="collapsed")

    # Check if any product has Wolt enrichment
    has_wolt = any(p.get("wolt_price") for p in products)

    # Table header
    if has_wolt:
        hcols = st.columns([3, 2, 1, 1, 1, 1])
        for col, label in zip(hcols, ["Product", "Barcode", "Qty", "Cost", "Sell", "Margin"]):
            col.markdown(f"**{label}**")
    else:
        hcols = st.columns([3, 2, 1, 1, 1])
        for col, label in zip(hcols, ["Product", "Barcode", "Qty", "Cost", "Status"]):
            col.markdown(f"**{label}**")
    st.markdown('<hr style="margin:4px 0">', unsafe_allow_html=True)

    shown = 0
    for i, p in enumerate(products):
        bc     = p.get("barcode", "")
        name   = p.get("name", "")
        dec    = decisions.get(i, {}).get("decision", "approve")
        status = dec

        if search and search.lower() not in name.lower() and search not in bc:
            continue
        if show == "Approved" and status != "approve":
            continue
        if show == "Pending"  and status != "pending":
            continue
        if show == "Skipped"  and status != "skip":
            continue

        badge_html = {
            "approve": '<span class="badge-approved">Approved</span>',
            "pending": '<span class="badge-pending">Pending</span>',
            "skip":    '<span class="badge-skipped">Skipped</span>',
        }.get(status, "")

        if has_wolt:
            row = st.columns([3, 2, 1, 1, 1, 1])
            row[0].write(name[:55] + ("…" if len(name) > 55 else ""))
            row[1].code(bc, language=None)
            row[2].write(p.get("quantity", ""))
            row[3].write(f"₪{p.get('cost', 0):.2f}")
            sell = p.get("wolt_price")
            row[4].write(f"₪{sell:.0f}" if sell else "—")
            margin = p.get("margin_pct")
            if margin is not None:
                color = "green" if margin >= 30 else "orange" if margin >= 15 else "red"
                row[5].markdown(f"<span style='color:{color};font-weight:500'>{margin:.0f}%</span>",
                                unsafe_allow_html=True)
            else:
                row[5].write("—")
        else:
            row = st.columns([3, 2, 1, 1, 1])
            row[0].write(name[:60] + ("…" if len(name) > 60 else ""))
            row[1].code(bc, language=None)
            row[2].write(p.get("quantity", ""))
            row[3].write(f"₪{p.get('cost', 0):.2f}")
            row[4].markdown(badge_html, unsafe_allow_html=True)
        shown += 1

    if shown == 0:
        st.caption("No products match the current filter.")

    st.markdown("---")
    if not st.session_state.saved:
        if st.button("Save to Google Sheets →", type="primary"):
            if st.session_state.pending:
                st.session_state.page = "review"
            else:
                with st.spinner("Saving…"):
                    result = save_to_sheets()
                st.session_state.save_result = result
                st.session_state.saved = True
            st.rerun()
    else:
        _post_save_actions()


# ─────────────────────────────────────────────
#  PAGE: CATALOG
# ─────────────────────────────────────────────

def page_catalog():
    st.title("Invoice Catalog")
    st.caption("All known products locked into the agent. New approvals are added here automatically.")

    try:
        from extract import PRODUCT_CATALOG
    except ImportError:
        st.error("Could not load PRODUCT_CATALOG from extract.py")
        return

    # Enrich with the shop catalog so we can show each product's buy price.
    try:
        from wolt_catalog import match_product
    except ImportError:
        match_product = None

    search = st.text_input("Search catalog", placeholder="Name or barcode…",
                           label_visibility="collapsed")

    st.markdown(f"**{len(PRODUCT_CATALOG)} products** in catalog")
    st.markdown("---")

    hcols = st.columns([2, 4, 1.4])
    hcols[0].markdown("**Barcode**")
    hcols[1].markdown("**Product name**")
    hcols[2].markdown("**Buy price ₪**")
    st.markdown('<hr style="margin:4px 0">', unsafe_allow_html=True)

    shown = 0
    for bc, name in PRODUCT_CATALOG.items():
        if search and search.lower() not in name.lower() and search not in bc:
            continue
        row = st.columns([2, 4, 1.4])
        row[0].code(bc, language=None)
        row[1].write(name)

        # Look the product up in the shop catalog (by barcode/SKU) → buy price.
        wolt = match_product(barcode=str(bc)) if match_product else None
        if wolt and wolt.get("wolt_id"):
            key = f"invbuy_{bc}"
            row[2].number_input(
                "Buy price", min_value=0.0, step=0.5,
                value=float(wolt.get("buy_price") or 0),
                key=key, on_change=_save_buy_price, args=(wolt["wolt_id"], key),
                label_visibility="collapsed",
            )
        else:
            row[2].caption("—")
        shown += 1

    if shown == 0:
        st.caption("No products match your search.")

    st.markdown("---")
    st.info(
        "To manually add a product, open `nodes/extract.py` and add an entry to `PRODUCT_CATALOG`. "
        "Products approved through the Review page are added automatically."
    )


# ─────────────────────────────────────────────
#  PAGE: WOLT CATALOG
# ─────────────────────────────────────────────

def _save_buy_price(pid, key):
    """on_change callback for a product's buy-price input."""
    from wolt_catalog import set_buy_price
    set_buy_price(pid, st.session_state.get(key, 0) or 0)


def page_wolt():
    st.markdown(
        '<h1 style="margin-bottom:0">🛒 קטלוג החנות</h1>'
        '<p style="color:#888;margin-top:2px">מחיר וולט · מחיר חנות · מחיר קנייה</p>',
        unsafe_allow_html=True,
    )
    # Mobile: show the product grid 2-per-row (instead of 1) by wrapping columns.
    # Scoped to this page — the <style> is only emitted when the catalog renders.
    st.markdown("""
    <style>
    @media (max-width: 768px) {
        section[data-testid="stMain"] div[data-testid="stHorizontalBlock"] {
            flex-wrap: wrap !important;
            gap: 0.5rem !important;
        }
        section[data-testid="stMain"] div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] {
            flex: 1 1 calc(50% - 0.5rem) !important;
            min-width: calc(50% - 0.5rem) !important;
            width: calc(50% - 0.5rem) !important;
        }
    }
    </style>
    """, unsafe_allow_html=True)

    try:
        from wolt_catalog import (
            is_loaded, all_products, categories,
            load_catalog_from_bytes, reload_catalog, FUZZY_THRESHOLD,
        )
        loaded = is_loaded()
    except ImportError:
        st.error("wolt_catalog.py not found — make sure it's in your nodes/ folder.")
        return

    if not loaded:
        st.info("עדיין לא נטען קטלוג. העלה קובץ Wolt/Emma בפאנל הניהול למטה.")

    # ── Catalog management (upload / reload / help) tucked away ───────────
    with st.expander("⚙️  ניהול קטלוג  ·  Manage catalog"):
        if loaded:
            from wolt_catalog import _barcode_index, _sku_index, _catalog
            st.success(f"✅ {len(_catalog)} מוצרים "
                       f"({len(_barcode_index)} עם ברקוד, {len(_sku_index)} עם מק\"ט)")
            if st.button("🔄 טען מחדש מהקובץ"):
                reload_catalog()
                st.rerun()

        uploaded = st.file_uploader("העלאת קטלוג (.xlsx)", type=["xlsx"],
                                    label_visibility="collapsed")
        if uploaded:
            c1, c2 = st.columns([3, 1])
            c1.info(f"**{uploaded.name}** — {uploaded.size // 1024} KB")
            if c2.button("טען", type="primary", use_container_width=True):
                with st.spinner("טוען…"):
                    ok = load_catalog_from_bytes(uploaded.read(), uploaded.name)
                if ok:
                    st.success("נטען!")
                    st.rerun()
                else:
                    st.error("טעינה נכשלה — ודא שזהו קובץ עם גיליון 'offers'.")

        st.caption(
            f"מחיר חנות = מחיר וולט −27%, מעוגל כלפי מעלה.  "
            f"התאמת חשבוניות: ברקוד → מק\"ט → שם ({int(FUZZY_THRESHOLD*100)}%+)."
        )

    if not loaded:
        return

    products = all_products()

    # ── Filter toolbar ───────────────────────────────────────────────────
    max_price = int(max((p.sell_price for p in products), default=0)) + 1
    f1, f2, f3 = st.columns([3, 2, 2])
    search = f1.text_input("חיפוש", placeholder="חפש לפי שם, ברקוד או מק\"ט…",
                           label_visibility="collapsed")
    cat_options = ["כל הקטגוריות"] + categories()
    chosen_cat = f2.selectbox("קטגוריה", cat_options, label_visibility="collapsed")
    PRICE_VIEWS = ["כל המחירים", "מחיר וולט", "מחיר חנות", "מחיר קנייה"]
    price_view = f3.selectbox("תצוגת מחיר", PRICE_VIEWS, label_visibility="collapsed")

    f4, f5, f6 = st.columns([3, 2, 2])
    price_lo, price_hi = f4.slider("טווח מחיר וולט (₪)", 0, max_price, (0, max_price))
    status = f5.selectbox("סטטוס", ["הכל", "פעילים בלבד", "לא פעילים בלבד"],
                          label_visibility="collapsed")
    missing_buy = f6.toggle("ללא מחיר קנייה", value=False)

    # ── Apply filters ────────────────────────────────────────────────────
    s = search.lower().strip()
    filtered = []
    for p in products:
        if s and s not in p.name.lower() and s not in p.barcode and s not in p.merchant_sku:
            continue
        if chosen_cat != "כל הקטגוריות" and p.category != chosen_cat:
            continue
        if not (price_lo <= p.sell_price <= price_hi):
            continue
        if status == "פעילים בלבד" and not p.enabled:
            continue
        if status == "לא פעילים בלבד" and p.enabled:
            continue
        if missing_buy and p.buy_price > 0:
            continue
        filtered.append(p)

    # ── Pagination ───────────────────────────────────────────────────────
    PER_PAGE = 24
    COLS = 4
    total = len(filtered)
    pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)

    top = st.columns([3, 1])
    top[0].caption(f"נמצאו **{total}** מוצרים")
    page = top[1].number_input("עמוד", 1, pages, 1, label_visibility="collapsed") if pages > 1 else 1
    start = (page - 1) * PER_PAGE
    page_items = filtered[start:start + PER_PAGE]

    if not page_items:
        st.info("לא נמצאו מוצרים התואמים את הסינון.")
        return

    # ── Product grid ─────────────────────────────────────────────────────
    for i in range(0, len(page_items), COLS):
        cols = st.columns(COLS)
        for col, p in zip(cols, page_items[i:i + COLS]):
            with col, st.container(border=True):
                if p.image_url:
                    st.image(p.thumb(300), use_container_width=True)
                else:
                    st.markdown("<div style='height:120px;text-align:center;"
                                "line-height:120px;color:#bbb'>🐾 אין תמונה</div>",
                                unsafe_allow_html=True)

                name = p.name if len(p.name) <= 60 else p.name[:58] + "…"
                st.markdown(f"**{name}**")
                chip = p.category + (f" · {p.subcategory}" if p.subcategory else "")
                st.caption(chip or "ללא קטגוריה")
                if not p.enabled:
                    st.caption("⚠️ לא פעיל")

                # ── Prices (🟦 Wolt / 🟩 Store / 🟨 Buy) — filtered by dropdown ──
                buy_txt = f"₪{p.buy_price:.0f}" if p.buy_price > 0 else "—"
                lines = []
                if price_view in ("כל המחירים", "מחיר וולט"):
                    lines.append(f"🟦 וולט <b>₪{p.sell_price:.0f}</b>")
                if price_view in ("כל המחירים", "מחיר חנות"):
                    lines.append(f"🟩 חנות <b>₪{p.store_price}</b>")
                if price_view in ("כל המחירים", "מחיר קנייה"):
                    lines.append(f"🟨 קנייה <b>{buy_txt}</b>")
                st.markdown(
                    "<div style='font-size:0.9em;line-height:1.6'>"
                    + "<br>".join(lines) + "</div>",
                    unsafe_allow_html=True,
                )

                # ── Action icons: ✏️ edit buy price · 📄 description ──
                a1, a2 = st.columns(2)
                buy_key = f"buy_{p.wolt_id}"
                with a1.popover("✏️", help="ערוך מחיר קנייה"):
                    st.number_input(
                        "מחיר קנייה ₪", min_value=0.0, step=0.5,
                        value=float(p.buy_price), key=buy_key,
                        on_change=_save_buy_price, args=(p.wolt_id, buy_key),
                    )
                with a2.popover("📄", help="תיאור ופרטים"):
                    st.markdown(f"**{p.name}**")
                    st.write(p.description if p.description else "_אין תיאור_")
                    meta = []
                    if p.weight_in_grams: meta.append(f"משקל: {p.weight_in_grams:.0f} גרם")
                    if p.volume_in_ml:    meta.append(f"נפח: {p.volume_in_ml:.0f} מ\"ל")
                    if p.number_of_units: meta.append(f"יחידות: {p.number_of_units:.0f}")
                    if p.barcode:         meta.append(f"ברקוד: {p.barcode}")
                    if p.merchant_sku:    meta.append(f"מק\"ט: {p.merchant_sku}")
                    if meta:
                        st.caption("  ·  ".join(meta))
                    if p.buy_price > 0 and p.store_price:
                        margin = (p.store_price - p.buy_price) / p.store_price * 100
                        st.caption(f"רווח בחנות: {margin:.0f}%")


# ─────────────────────────────────────────────
#  PAGE: IMAGE PROCESSOR
# ─────────────────────────────────────────────

class _ExternalImage:
    """Adapter that mimics Streamlit's UploadedFile for images pushed in via session state."""
    def __init__(self, name, data):
        self.name = name
        self._data = data
    def getvalue(self):
        return self._data


def page_image_processor():
    st.title("Image Processor")
    st.caption("Upload one or more product images to get 1000×1000 and 1000×563 PNGs with background removed.")

    uploaded_files = st.file_uploader(
        "Upload product images",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    # Images queued from other pages (e.g. Product Extractor's "Send to Image Processor")
    queued = st.session_state.get("_imgproc_external", [])
    if queued:
        c1, c2 = st.columns([4, 1])
        with c1:
            st.info(f"📥 {len(queued)} image(s) queued from Product Extractor.")
        with c2:
            if st.button("Clear queue", use_container_width=True):
                st.session_state["_imgproc_external"] = []
                st.rerun()

    external_items = [_ExternalImage(e["name"], e["bytes"]) for e in queued]
    all_files = external_items + list(uploaded_files or [])

    # Clear stale results when the working set changes
    file_key = tuple(f.name for f in all_files)
    if st.session_state.get("_imgproc_files") != file_key:
        st.session_state["_imgproc_files"] = file_key
        st.session_state["_imgproc_batch"] = []

    skip_rembg = st.checkbox("Background already removed — skip background removal", value=False)
    add_shadow = st.checkbox("Add drop shadow", value=True)

    uploaded_files = all_files  # use the merged list for the rest of the page
    if uploaded_files:
        import io as _io
        from PIL import Image

        st.caption(f"{len(uploaded_files)} image(s) selected — set rotation per image if needed; output names are derived from each filename.")

        # Per-image rotation grid (4 thumbs per row)
        THUMBS_PER_ROW = 4
        rot_options = [0, 90, 180, 270]
        for row_start in range(0, len(uploaded_files), THUMBS_PER_ROW):
            cols = st.columns(THUMBS_PER_ROW)
            for col, uploaded in zip(cols, uploaded_files[row_start:row_start + THUMBS_PER_ROW]):
                with col:
                    rot = st.session_state.get(f"rot_{uploaded.name}", 0)
                    try:
                        preview = Image.open(_io.BytesIO(uploaded.getvalue()))
                        if rot:
                            preview = preview.rotate(-rot, expand=True)
                        st.image(preview, caption=uploaded.name, use_container_width=True)
                    except Exception as e:
                        st.error(f"⚠️ {uploaded.name}: not a valid image ({e})")
                    st.selectbox(
                        "Rotate",
                        options=rot_options,
                        format_func=lambda x: "0°" if x == 0 else f"{x}° CW",
                        key=f"rot_{uploaded.name}",
                        label_visibility="collapsed",
                    )

        if st.button(f"Process {len(uploaded_files)} image(s)", type="primary", use_container_width=True):
            try:
                from rembg import remove
            except Exception:
                remove = None
            from tools.image_utils.processor import create_formatted_image, sanitize_filename
            import numpy as np

            if remove is None and not skip_rembg:
                st.warning(
                    "Background removal isn't available on this server (rembg not installed). "
                    "Images will be formatted **without** background removal — upload "
                    "already-transparent PNGs, tick “Background already removed”, or run "
                    "locally for full background removal."
                )

            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_images")
            os.makedirs(output_dir, exist_ok=True)
            batch = []
            progress = st.progress(0.0, text="Starting…")

            for idx, uploaded in enumerate(uploaded_files, start=1):
                progress.progress((idx - 1) / len(uploaded_files), text=f"[{idx}/{len(uploaded_files)}] {uploaded.name}")
                stem = uploaded.name.rsplit(".", 1)[0].replace("_", " ").replace("-", " ").strip()
                safe_name = sanitize_filename(stem)
                raw = uploaded.getvalue()
                rot = st.session_state.get(f"rot_{uploaded.name}", 0)

                # Apply rotation up-front so rembg + downstream all see the rotated image
                if rot:
                    rotated = Image.open(_io.BytesIO(raw)).rotate(-rot, expand=True)
                    rot_buf = _io.BytesIO()
                    rotated.save(rot_buf, format="PNG")
                    raw = rot_buf.getvalue()

                try:
                    img = Image.open(_io.BytesIO(raw)).convert("RGBA")
                    arr = np.array(img)

                    total_px = arr.shape[0] * arr.shape[1]
                    already_transparent = np.sum(arr[:, :, 3] == 0) / total_px > 0.10

                    if not (skip_rembg or already_transparent or remove is None):
                        arr = np.array(Image.open(_io.BytesIO(remove(raw))).convert("RGBA"))

                    arr[arr[:, :, 3] < 20, 3] = 0
                    img = Image.fromarray(arr)
                    bbox = img.getbbox()
                    if not bbox:
                        batch.append({"src": uploaded.name, "error": "No product detected", "variants": []})
                        continue

                    cropped = img.crop(bbox)
                    variants = [
                        ((1000, 1000), "1000×1000", f"{safe_name}.png"),
                        ((1000, 563),  "1000×563",  f"wolt{safe_name}.png"),
                    ]
                    variant_results = []
                    for size, tag, fname in variants:
                        out = create_formatted_image(cropped, size, drop_shadow=add_shadow)
                        out.save(os.path.join(output_dir, fname), "PNG")
                        buf = _io.BytesIO()
                        out.save(buf, "PNG")
                        variant_results.append((buf.getvalue(), fname, tag))
                    batch.append({"src": uploaded.name, "error": None, "variants": variant_results})
                except Exception as e:
                    batch.append({"src": uploaded.name, "error": str(e), "variants": []})

            progress.progress(1.0, text="Done!")
            st.session_state["_imgproc_batch"] = batch

    else:
        st.markdown("""
        <div style="text-align:center;padding:3rem 1rem">
            <p style="font-size:1.1rem;font-weight:600;color:#444;margin:0.5rem 0">No images yet</p>
            <p style="color:#999;font-size:0.9rem">Drop one or more product photos above to get started.</p>
        </div>""", unsafe_allow_html=True)

    # Results — rendered outside the button block so they survive reruns from download clicks
    batch = st.session_state.get("_imgproc_batch", [])
    if batch:
        st.markdown("---")

        # Bundle all variants into a single zip for convenience
        import io as _io
        import zipfile as _zip
        zip_buf = _io.BytesIO()
        with _zip.ZipFile(zip_buf, "w", _zip.ZIP_DEFLATED) as zf:
            for item in batch:
                for data, fname, _tag in item["variants"]:
                    zf.writestr(fname, data)
        if zip_buf.getbuffer().nbytes > 0:
            st.download_button(
                "Download all (.zip)",
                data=zip_buf.getvalue(),
                file_name="processed_images.zip",
                mime="application/zip",
                use_container_width=True,
                key="dl_all_zip",
            )

        for i, item in enumerate(batch):
            st.markdown(f"**{item['src']}**")
            if item["error"]:
                st.error(item["error"])
                continue
            cols = st.columns([1, 1, 2])  # two small previews on the left, empty space on the right
            for col, (data, fname, tag) in zip(cols[:2], item["variants"]):
                with col:
                    st.image(data, caption=tag, width=200)
                    st.download_button(
                        "Download",
                        data=data,
                        file_name=fname,
                        mime="image/png",
                        use_container_width=True,
                        key=f"dl_{i}_{tag}",
                    )


# ─────────────────────────────────────────────
#  PAGE: HOME
# ─────────────────────────────────────────────

def page_home():
    st.markdown("## Invoice Agent")
    st.caption("AI-powered invoice processing and product image automation.")
    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div style="border:1px solid #e0e0e0;border-radius:12px;padding:1.5rem 1rem;text-align:center;min-height:130px">
            <svg width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="#888" stroke-width="1.6"
                 stroke-linecap="round" stroke-linejoin="round" style="margin-bottom:8px">
                <path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/>
                <polyline points="14 2 14 8 20 8"/>
                <line x1="12" y1="18" x2="12" y2="12"/>
                <line x1="9" y1="15" x2="15" y2="15"/>
            </svg>
            <p style="font-weight:600;margin:0 0 4px">Upload Invoice</p>
            <p style="font-size:0.8rem;color:#999;margin:0">Extract products from PDF or image invoices</p>
        </div>""", unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("Open", key="home_upload", use_container_width=True, type="primary"):
            st.session_state.page = "upload"
            st.rerun()

    with col2:
        st.markdown("""
        <div style="border:1px solid #e0e0e0;border-radius:12px;padding:1.5rem 1rem;text-align:center;min-height:130px">
            <svg width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="#888" stroke-width="1.6"
                 stroke-linecap="round" stroke-linejoin="round" style="margin-bottom:8px">
                <circle cx="9" cy="21" r="1"/>
                <circle cx="20" cy="21" r="1"/>
                <path d="M1 1h4l2.68 13.39a2 2 0 002 1.61h9.72a2 2 0 002-1.61L23 6H6"/>
            </svg>
            <p style="font-weight:600;margin:0 0 4px">Catalog</p>
            <p style="font-size:0.8rem;color:#999;margin:0">Browse products, prices & categories</p>
        </div>""", unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("Open", key="home_catalog", use_container_width=True):
            st.session_state.page = "wolt"
            st.rerun()

    with col3:
        st.markdown("""
        <div style="border:1px solid #e0e0e0;border-radius:12px;padding:1.5rem 1rem;text-align:center;min-height:130px">
            <svg width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="#888" stroke-width="1.6"
                 stroke-linecap="round" stroke-linejoin="round" style="margin-bottom:8px">
                <rect x="3" y="3" width="18" height="18" rx="2" ry="2"/>
                <circle cx="8.5" cy="8.5" r="1.5"/>
                <polyline points="21 15 16 10 5 21"/>
            </svg>
            <p style="font-weight:600;margin:0 0 4px">Process Image</p>
            <p style="font-size:0.8rem;color:#999;margin:0">Remove backgrounds & create catalog images</p>
        </div>""", unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("Open", key="home_images", use_container_width=True):
            st.session_state.page = "image_processor"
            st.rerun()

    if st.session_state.products:
        st.markdown("---")
        st.caption("Current session")
        c1, c2, c3, c4 = st.columns(4)
        total_cost = sum(p.get("cost", 0) for p in st.session_state.approved)
        c1.metric("Supplier",  st.session_state.vendor_name or "—")
        c2.metric("Products",  len(st.session_state.products))
        c3.metric("Pending",   len(st.session_state.pending))
        c4.metric("Total",     f"₪{total_cost:,.0f}")
        if st.button("View results →", use_container_width=True):
            st.session_state.page = "results"
            st.rerun()


# ─────────────────────────────────────────────
#  PAGE: WOLT EXPORT
# ─────────────────────────────────────────────

def page_wolt_export():
    import pandas as pd
    import zipfile
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    st.title("Wolt Export")
    st.caption("Build the Wolt product Excel, bundle catalog images, and submit via Monday.")

    # ── Section 1: Product table ──────────────────────────────────────────
    st.subheader("Product table")

    HEBREW_HEADERS = ["שם פריט", "מק״ט", "תיאור פריט", "מחיר פריט",
                      "משקל פריט", "נפח פריט", "הערות"]
    COL_KEYS       = ["שם פריט", "מק״ט", "תיאור פריט", "מחיר פריט",
                      "משקל פריט", "נפח פריט", "הערות"]

    # Pre-populate from session if available
    products = st.session_state.get("products", [])
    decisions = st.session_state.get("decisions", {})
    approved_products = [
        p for i, p in enumerate(products)
        if decisions.get(i, {}).get("decision", "approve") == "approve"
    ]

    if approved_products:
        st.caption(f"Pre-filled from current session ({len(approved_products)} approved products). Edit any cell before exporting.")
        rows = []
        for p in approved_products:
            rows.append({
                "שם פריט":       p.get("name", ""),
                "מק״ט":          p.get("sku") or p.get("wolt_sku") or "",
                "תיאור פריט":    "",
                "מחיר פריט":     p.get("wolt_price") or None,
                "משקל פריט":     None,
                "נפח פריט":      None,
                "הערות":         "",
            })
        df = pd.DataFrame(rows, columns=COL_KEYS)
    else:
        st.caption("No invoice loaded. Fill in the table manually or upload an invoice first.")
        df = pd.DataFrame(
            [{"שם פריט": "", "מק״ט": "", "תיאור פריט": "", "מחיר פריט": None,
              "משקל פריט": None, "נפח פריט": None, "הערות": ""}],
            columns=COL_KEYS,
        )

    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "שם פריט":       st.column_config.TextColumn("שם פריט", width="large"),
            "מק״ט":          st.column_config.TextColumn("מק״ט", width="small"),
            "תיאור פריט":    st.column_config.TextColumn("תיאור פריט", width="medium"),
            "מחיר פריט":     st.column_config.NumberColumn("מחיר פריט ₪", format="%.2f"),
            "משקל פריט":     st.column_config.NumberColumn("משקל (ג׳)", format="%.0f"),
            "נפח פריט":      st.column_config.NumberColumn("נפח (מ״ל)", format="%.0f"),
            "הערות":         st.column_config.TextColumn("הערות"),
        },
        key="wolt_export_table",
    )

    # ── Section 2: Generate Excel ─────────────────────────────────────────
    st.markdown("---")
    st.subheader("Download Excel")

    if st.button("Generate Excel", type="primary", use_container_width=True):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(HEBREW_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align

        ws.row_dimensions[1].height = 22

        for row_idx, row in enumerate(edited_df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None and val != "" else None)
                cell.alignment = Alignment(horizontal="right")

        col_widths = [35, 14, 30, 12, 12, 12, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        vendor = st.session_state.get("vendor_name", "products")
        fname  = f"wolt_export_{vendor}.xlsx".replace(" ", "_")
        st.download_button(
            "⬇ Download Excel",
            data=buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success(f"Ready — {len(edited_df)} products in {fname}")

    # ── Section 3: Bundle images ──────────────────────────────────────────
    st.markdown("---")
    st.subheader("Bundle product images")
    st.caption(
        "Images processed via the Image Processor are listed below. "
        "Name your product images to match column A of the Excel so Wolt can pair them automatically."
    )

    img_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_images")
    png_files = sorted(f for f in os.listdir(img_dir) if f.endswith(".png")) if os.path.isdir(img_dir) else []

    if png_files:
        selected = st.multiselect("Select images to include in ZIP", png_files, default=png_files[:min(len(png_files), 10)])
        if selected and st.button("Download Images ZIP", use_container_width=True):
            zip_buf = _io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname in selected:
                    zf.write(os.path.join(img_dir, fname), fname)
            zip_buf.seek(0)
            st.download_button(
                "⬇ Download ZIP",
                data=zip_buf,
                file_name="wolt_images.zip",
                mime="application/zip",
                use_container_width=True,
                key="dl_zip",
            )
    else:
        st.info("No processed images yet — use the Image Processor page to create catalog images first.")

    # ── Section 4: Submit ─────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Submit to Wolt")
    st.markdown("""
1. Download the Excel file above
2. Download the images ZIP
3. Open the Monday form and attach both files
""")
    st.link_button(
        "Open Monday Form →",
        "https://forms.monday.com/forms/2386872fef915d70f6a9ba86cea156f6?r=use1",
        use_container_width=True,
        type="primary",
    )


# ─────────────────────────────────────────────
#  PAGE: PRODUCT PAGE EXTRACTOR
# ─────────────────────────────────────────────

def _extract_product_from_html(html: str, base_url: str = ""):
    """Pull (image_url, name, description) from a product page's HTML.

    Strategy: JSON-LD Product schema → Open Graph → DOM fallbacks.
    """
    import json
    import re as _re
    import unicodedata
    from urllib.parse import urljoin, urlparse
    from bs4 import BeautifulSoup

    def _tokens(s: str):
        """Lowercase, strip accents, split on non-alphanum. Drops short/stop tokens."""
        if not s:
            return set()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        s = s.replace("æ", "ae").replace("Æ", "ae")
        parts = _re.split(r"[^a-z0-9]+", s.lower())
        STOP = {"the", "and", "for", "with", "of", "a", "an", "in", "on", "to", "g", "kg", "ml"}
        return {p for p in parts if len(p) >= 3 and p not in STOP}

    soup = BeautifulSoup(html, "html.parser")
    name = ""
    desc = ""
    image = ""

    # 1) JSON-LD schema.org/Product
    for tag in soup.find_all("script", {"type": "application/ld+json"}):
        try:
            data = json.loads(tag.string or "")
        except Exception:
            continue
        candidates = data if isinstance(data, list) else [data]
        # @graph wrapper (common on WP/Yoast)
        if isinstance(data, dict) and "@graph" in data:
            candidates = data["@graph"]
        for c in candidates:
            if not isinstance(c, dict):
                continue
            t = c.get("@type")
            if t == "Product" or (isinstance(t, list) and "Product" in t):
                name  = name  or (c.get("name") or "")
                desc  = desc  or (c.get("description") or "")
                img_v = c.get("image")
                if isinstance(img_v, list):
                    img_v = img_v[0] if img_v else ""
                if isinstance(img_v, dict):
                    img_v = img_v.get("url", "")
                image = image or (img_v or "")

    # 2) Open Graph
    def _og(prop):
        m = soup.find("meta", attrs={"property": prop})
        return (m.get("content") or "").strip() if m else ""

    name  = name  or _og("og:title")
    desc  = desc  or _og("og:description")
    image = image or _og("og:image")

    # 3) DOM fallbacks
    if not name:
        h1 = soup.find("h1")
        if h1:
            name = h1.get_text(strip=True)
    if not desc:
        md = soup.find("meta", attrs={"name": "description"})
        if md and md.get("content"):
            desc = md["content"].strip()
    if not desc:
        # WooCommerce typical containers
        for sel in [".woocommerce-product-details__short-description", "#tab-description", ".product-description"]:
            el = soup.select_one(sel)
            if el:
                desc = el.get_text(" ", strip=True)
                break

    # Build a token set from the product name + URL slug so we can score
    # generic images / paragraphs against the product when no semantic markup exists.
    slug = ""
    if base_url:
        path = urlparse(base_url).path
        slug = path.rstrip("/").rsplit("/", 1)[-1]
    product_tokens = _tokens(name) | _tokens(slug)

    if not image:
        # WooCommerce gallery first
        gallery = soup.select_one(".woocommerce-product-gallery__image")
        if gallery:
            a = gallery.find("a")
            img = gallery.find("img")
            if a and a.get("href"):
                image = a["href"]
            elif img:
                image = img.get("data-large_image") or img.get("data-src") or img.get("src") or ""
        if not image:
            other = soup.select_one(".product-gallery img, .product-images img, .wp-post-image")
            if other:
                image = other.get("data-large_image") or other.get("data-src") or other.get("src") or ""

    if not image and product_tokens:
        # Generic scoring: pick the <img> whose alt+src share the most tokens with the product
        BAD = ("logo", "icon", "favicon", "placeholder", "spacer", "tracker",
               "facebook.com/tr", "google", "/symboli", "/glutenfree", "/monoproteico")
        best = (0, "")
        for img in soup.find_all("img"):
            src = img.get("data-src") or img.get("data-original") or img.get("src") or ""
            if not src or src.startswith("data:"):
                continue
            low = src.lower()
            if any(b in low for b in BAD):
                continue
            haystack = (img.get("alt", "") + " " + src).lower()
            score = sum(1 for t in product_tokens if t in haystack)
            if score > best[0]:
                best = (score, src)
        if best[0] > 0:
            image = best[1]

    if not desc:
        # Longest meaningful paragraph in the body, excluding nav/header/footer
        candidates = []
        for p in soup.find_all(["p", "div"]):
            if p.find_parent(["nav", "header", "footer", "aside"]):
                continue
            cls = " ".join(p.get("class", []) or []).lower()
            if any(b in cls for b in ("nav", "menu", "footer", "header", "cookie", "newsletter")):
                continue
            text = p.get_text(" ", strip=True)
            if 80 <= len(text) <= 2000:
                candidates.append(text)
        if candidates:
            desc = max(candidates, key=len)

    if image and base_url:
        image = urljoin(base_url, image)

    return image, (name or "").strip(), (desc or "").strip()


def page_product_extract():
    import requests

    st.title("Product Page Extractor")
    st.caption("Paste a product URL, drop an .html file, or paste HTML — pull out the image, name and description.")

    mode = st.radio(
        "Source",
        ["URL", "Search", "HTML file", "Paste HTML"],
        horizontal=True,
        label_visibility="collapsed",
    )

    html = ""
    base_url = ""

    if mode == "URL":
        # Quick picker — supplier or one of their brand sites
        merged = get_merged_suppliers()
        quick_opts = [("—", "")]
        for _, p in merged.items():
            if p.get("website"):
                quick_opts.append((f"{p['display_name']}", p["website"]))
            for b in p.get("brands", []) or []:
                if b.get("website"):
                    quick_opts.append((f"{p['display_name']} → {b['name']}", b["website"]))

        if len(quick_opts) > 1:
            labels = [lbl for lbl, _ in quick_opts]
            picked = st.selectbox("Quick start from supplier / brand (optional)", labels, index=0)
            picked_url = dict(quick_opts).get(picked, "")
            if picked != "—" and st.session_state.get("_prodext_supplier") != picked:
                st.session_state["_prodext_supplier"] = picked
                st.session_state["_prodext_url"] = picked_url

        url = st.text_input("Product URL", key="_prodext_url", placeholder="https://dudi-agencies.co.il/product/...")
        if st.button("Fetch & extract", type="primary", use_container_width=True) and url:
            try:
                resp = requests.get(
                    url,
                    headers={"User-Agent": "Mozilla/5.0 (compatible; EmmaAgent/1.0)"},
                    timeout=15,
                )
                resp.raise_for_status()
                html = resp.text
                base_url = url
            except Exception as e:
                st.error(f"Fetch failed: {e}")

    elif mode == "Search":
        from urllib.parse import urljoin, urlencode
        from bs4 import BeautifulSoup

        merged = get_merged_suppliers()
        site_opts = [("—", "")]
        for _, p in merged.items():
            if p.get("website"):
                site_opts.append((p["display_name"], p["website"]))
            for b in p.get("brands", []) or []:
                if b.get("website"):
                    site_opts.append((f"{p['display_name']} → {b['name']}", b["website"]))

        if len(site_opts) < 2:
            st.info("Add a supplier with a website on the Suppliers page first.")
        else:
            labels = [lbl for lbl, _ in site_opts]
            picked = st.selectbox("Site to search", labels, index=0)
            site_url = dict(site_opts).get(picked, "")
            query = st.text_input("Barcode or product name", placeholder="e.g. 'gosbi puppy' or 8430235681255")
            st.caption("Most sites only index product names (not barcodes) for search.")

            if st.button("Search", type="primary", use_container_width=True) and site_url and query:
                search_url = urljoin(site_url, "/?" + urlencode({"s": query, "post_type": "product"}))
                try:
                    r = requests.get(
                        search_url,
                        headers={"User-Agent": "Mozilla/5.0 (compatible; EmmaAgent/1.0)"},
                        timeout=15,
                    )
                    r.raise_for_status()
                    soup = BeautifulSoup(r.text, "html.parser")

                    results = []
                    seen = set()
                    for a in soup.find_all("a", href=True):
                        href = urljoin(site_url, a["href"])
                        low = href.lower()
                        if "/product/" not in low and "/products/" not in low:
                            continue
                        if href in seen:
                            continue
                        seen.add(href)
                        # Find a meaningful image in the result card by walking up the DOM
                        img_tag = None
                        parent = a
                        for _ in range(4):
                            if parent is None:
                                break
                            cand = parent.find("img")
                            if cand:
                                src = (cand.get("data-src") or cand.get("src") or "").lower()
                                if src and "logo" not in src and "icon" not in src and not src.startswith("data:"):
                                    img_tag = cand
                                    break
                            parent = parent.parent
                        title = a.get_text(" ", strip=True) or (img_tag.get("alt", "") if img_tag else "")
                        img_src = ""
                        if img_tag:
                            img_src = img_tag.get("data-src") or img_tag.get("src") or ""
                            if img_src:
                                img_src = urljoin(site_url, img_src)
                        results.append({"url": href, "title": title.strip(), "image": img_src})
                        if len(results) >= 24:
                            break

                    st.session_state["_prodext_search_results"] = results
                    st.session_state["_prodext_search_meta"] = {"site": picked, "query": query}
                except Exception as e:
                    st.error(f"Search failed: {e}")

            results = st.session_state.get("_prodext_search_results", [])
            meta = st.session_state.get("_prodext_search_meta", {})
            if results:
                st.markdown("---")
                st.caption(f"{len(results)} result(s) on {meta.get('site','')} for '{meta.get('query','')}'.")
                cols = st.columns(3)
                for i, rr in enumerate(results):
                    with cols[i % 3]:
                        if rr["image"]:
                            st.image(rr["image"], width=180)
                        st.caption((rr["title"] or "(no title)")[:80])
                        if st.button("Extract", key=f"_sext_{i}", use_container_width=True):
                            try:
                                page = requests.get(
                                    rr["url"],
                                    headers={"User-Agent": "Mozilla/5.0 (compatible; EmmaAgent/1.0)"},
                                    timeout=15,
                                )
                                page.raise_for_status()
                                image_url, name, desc = _extract_product_from_html(page.text, base_url=rr["url"])
                                st.session_state["_prodext_result"] = {
                                    "image_url": image_url, "name": name, "description": desc,
                                }
                                st.session_state["_prodext_name"] = name
                                st.session_state["_prodext_desc"] = desc
                                st.rerun()
                            except Exception as e:
                                st.error(f"Fetch failed: {e}")
            elif st.session_state.get("_prodext_search_meta"):
                st.info(
                    f"No products found on {meta.get('site','')} for '{meta.get('query','')}'. "
                    "Try a product-name keyword (most sites don't index barcodes)."
                )

    elif mode == "HTML file":
        uploaded = st.file_uploader("HTML file", type=["html", "htm"], label_visibility="collapsed")
        if uploaded and st.button("Extract", type="primary", use_container_width=True):
            html = uploaded.getvalue().decode("utf-8", errors="replace")

    else:  # Paste HTML
        pasted = st.text_area("Paste HTML", height=200, label_visibility="collapsed")
        page_url = st.text_input("(Optional) source URL — used to resolve relative image links")
        if st.button("Extract", type="primary", use_container_width=True) and pasted:
            html = pasted
            base_url = page_url

    if html:
        image_url, name, desc = _extract_product_from_html(html, base_url=base_url)
        st.session_state["_prodext_result"] = {
            "image_url": image_url,
            "name": name,
            "description": desc,
        }
        # Sync editable fields to the new extraction (passing value= alongside key= would be ignored on rerun)
        st.session_state["_prodext_name"] = name
        st.session_state["_prodext_desc"] = desc

    result = st.session_state.get("_prodext_result")
    if result:
        st.markdown("---")
        left, right = st.columns([1, 2])
        with left:
            if result["image_url"]:
                st.image(result["image_url"], caption="Product image", width=240)
                st.caption(result["image_url"])
            else:
                st.info("No image found.")
        with right:
            st.text_input("Name", key="_prodext_name")
            st.text_area("Description", height=180, key="_prodext_desc")

        if result["image_url"]:
            try:
                img_resp = requests.get(
                    result["image_url"],
                    headers={"User-Agent": "Mozilla/5.0 (compatible; EmmaAgent/1.0)"},
                    timeout=15,
                )
                if img_resp.ok:
                    ext = result["image_url"].rsplit(".", 1)[-1].lower().split("?")[0]
                    if ext not in ("jpg", "jpeg", "png", "webp"):
                        ext = "jpg"
                    safe_name = (st.session_state.get("_prodext_name") or "product").strip() or "product"
                    fname = f"{safe_name}.{ext}"
                    mime  = f"image/{ 'jpeg' if ext == 'jpg' else ext }"

                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.download_button(
                            "Download image",
                            data=img_resp.content,
                            file_name=fname,
                            mime=mime,
                            use_container_width=True,
                        )
                    with col_b:
                        if st.button("Send to Image Processor →", type="primary", use_container_width=True):
                            import io as _io
                            from PIL import Image as _PIL
                            ctype = img_resp.headers.get("content-type", "").lower()
                            if "svg" in ctype or result["image_url"].lower().endswith(".svg"):
                                st.error("Image is SVG, which can't be processed. Try a different product photo.")
                            else:
                                try:
                                    _PIL.open(_io.BytesIO(img_resp.content)).verify()
                                except Exception:
                                    st.error(
                                        f"The URL didn't return a usable image "
                                        f"(content-type: {ctype or 'unknown'}). "
                                        f"It may be a redirect, an HTML page, or an unsupported format."
                                    )
                                else:
                                    queue = st.session_state.setdefault("_imgproc_external", [])
                                    queue.append({"name": fname, "bytes": img_resp.content})
                                    st.session_state.page = "image_processor"
                                    st.rerun()
            except Exception as e:
                st.caption(f"(Image download unavailable: {e})")


# ─────────────────────────────────────────────
#  PAGE: SUPPLIERS DIRECTORY
# ─────────────────────────────────────────────

SUPPLIER_OVERRIDES_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "supplier_overrides.json"
)


def _load_supplier_overrides() -> dict:
    import json
    if not os.path.exists(SUPPLIER_OVERRIDES_PATH):
        return {}
    try:
        with open(SUPPLIER_OVERRIDES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_supplier_overrides(data: dict):
    import json
    with open(SUPPLIER_OVERRIDES_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_merged_suppliers() -> dict:
    """Merge SUPPLIER_PROFILES defaults with the editable overrides JSON.
    Overrides can edit `website`/`brands` for built-in suppliers,
    or add entirely new suppliers via `_added: true` entries (display_name/website/brands only —
    no invoice-extraction profile, so they won't auto-match on invoice header recognition)."""
    from nodes.suppliers import SUPPLIER_PROFILES
    overrides = _load_supplier_overrides()
    merged = {}
    for key, p in SUPPLIER_PROFILES.items():
        m = dict(p)
        ov = overrides.get(key, {})
        if "website" in ov:
            m["website"] = ov["website"]
        if "brands" in ov:
            m["brands"] = ov["brands"]
        m.setdefault("brands", [])
        merged[key] = m
    # Append user-added suppliers (those in overrides but not in defaults)
    for key, ov in overrides.items():
        if key in merged:
            continue
        merged[key] = {
            "display_name": ov.get("display_name", key),
            "website":      ov.get("website", ""),
            "brands":       ov.get("brands", []),
            "id_patterns":  [],
            "_added":       True,
        }
    return merged


def _slugify_key(name: str) -> str:
    import re as _re
    import unicodedata
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = _re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s or "supplier"


def page_suppliers():
    import pandas as pd

    st.title("Suppliers")
    st.caption("Known suppliers and the brands they distribute. Edits persist to `supplier_overrides.json`.")

    suppliers = get_merged_suppliers()

    rows = []
    for key, p in suppliers.items():
        brand_str = ", ".join(b["name"] for b in p.get("brands", []) if b.get("name")) or "—"
        rows.append({
            "Name":    p["display_name"] + (" ✎" if p.get("_added") else ""),
            "Website": p.get("website") or "",
            "Brands":  brand_str,
        })
    df = pd.DataFrame(rows)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Website": st.column_config.LinkColumn("Website"),
            "Brands":  st.column_config.TextColumn("Brands", width="large"),
        },
    )

    # ── Edit supplier ────────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("✏️ Edit supplier"):
        key_to_name = {k: p["display_name"] for k, p in suppliers.items()}
        selected_key = st.selectbox(
            "Supplier",
            options=list(key_to_name.keys()),
            format_func=lambda k: key_to_name[k],
            key="_sup_edit_key",
        )
        cur = suppliers[selected_key]

        website = st.text_input("Website", value=cur.get("website", ""), key=f"_sup_web_{selected_key}")

        st.caption("Brands this supplier distributes")
        brands_df = pd.DataFrame(cur.get("brands", []) or [], columns=["name", "website"])
        if brands_df.empty:
            brands_df = pd.DataFrame([{"name": "", "website": ""}])
        edited_brands = st.data_editor(
            brands_df,
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "name":    st.column_config.TextColumn("Brand name"),
                "website": st.column_config.TextColumn("Brand website"),
            },
            key=f"_sup_brands_{selected_key}",
        )

        col_a, col_b, col_c = st.columns([1, 1, 4])
        with col_a:
            if st.button("Save", type="primary"):
                overrides = _load_supplier_overrides()
                cleaned_brands = [
                    {"name": str(r.get("name", "")).strip(), "website": str(r.get("website", "")).strip()}
                    for _, r in edited_brands.iterrows()
                    if str(r.get("name", "")).strip()
                ]
                existing = overrides.get(selected_key, {})
                existing["website"] = website.strip()
                existing["brands"]  = cleaned_brands
                if cur.get("_added"):
                    existing["display_name"] = cur["display_name"]
                    existing["_added"] = True
                overrides[selected_key] = existing
                _save_supplier_overrides(overrides)
                st.success(f"Saved {cur['display_name']}.")
                st.rerun()
        with col_b:
            if cur.get("_added") and st.button("Delete", type="secondary"):
                overrides = _load_supplier_overrides()
                overrides.pop(selected_key, None)
                _save_supplier_overrides(overrides)
                st.success(f"Deleted {cur['display_name']}.")
                st.rerun()
        with col_c:
            st.caption(f"Override file: `{os.path.basename(SUPPLIER_OVERRIDES_PATH)}`")

    # ── Add new supplier ─────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("➕ Add new supplier"):
        st.caption(
            "User-added suppliers appear in the directory and Product Extractor picker, "
            "but won't be auto-recognized in invoice extraction "
            "(that requires editing `nodes/suppliers.py`)."
        )
        new_name = st.text_input("Display name", key="_sup_new_name", placeholder="e.g. ACME Pet Distributors")
        new_website = st.text_input("Website", key="_sup_new_website", placeholder="https://...")

        new_brands_df = pd.DataFrame([{"name": "", "website": ""}])
        new_edited_brands = st.data_editor(
            new_brands_df,
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "name":    st.column_config.TextColumn("Brand name"),
                "website": st.column_config.TextColumn("Brand website"),
            },
            key="_sup_new_brands",
        )

        if st.button("Add supplier", type="primary"):
            name = (new_name or "").strip()
            if not name:
                st.error("Display name is required.")
            else:
                overrides = _load_supplier_overrides()
                base_key = _slugify_key(name)
                key = base_key
                i = 2
                # Avoid colliding with built-in keys or existing overrides
                taken = set(suppliers.keys()) | set(overrides.keys())
                while key in taken:
                    key = f"{base_key}_{i}"
                    i += 1
                cleaned = [
                    {"name": str(r.get("name", "")).strip(), "website": str(r.get("website", "")).strip()}
                    for _, r in new_edited_brands.iterrows()
                    if str(r.get("name", "")).strip()
                ]
                overrides[key] = {
                    "display_name": name,
                    "website": (new_website or "").strip(),
                    "brands":  cleaned,
                    "_added":  True,
                }
                _save_supplier_overrides(overrides)
                st.success(f"Added {name} (key: `{key}`).")
                st.rerun()


# ─────────────────────────────────────────────
#  MOBILE BOTTOM NAV
# ─────────────────────────────────────────────

def mobile_bottom_nav():
    current = st.session_state.page

    SVG = {
        "home": '<svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9l9-7 9 7v11a2 2 0 01-2 2H5a2 2 0 01-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>',
        "upload": '<svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>',
        "review": '<svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>',
        "results": '<svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/><line x1="8" y1="18" x2="21" y2="18"/><line x1="3" y1="6" x2="3.01" y2="6"/><line x1="3" y1="12" x2="3.01" y2="12"/><line x1="3" y1="18" x2="3.01" y2="18"/></svg>',
        "images": '<svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"/><circle cx="8.5" cy="8.5" r="1.5"/><polyline points="21 15 16 10 5 21"/></svg>',
    }

    items = [
        (SVG["home"],    "Home",    "home",            "Home"),
        (SVG["upload"],  "Upload",  "upload",          "Upload"),
    ]
    if st.session_state.products or current in ("review", "results"):
        items += [
            (SVG["review"],  "Review",  "review",  "Review"),
            (SVG["results"], "Results", "results", "Results"),
        ]
    items.append((SVG["images"], "Images", "image_processor", "Image"))

    nav_items_html = ""
    for svg, label, page, sidebar_text in items:
        active = "active" if current == page else ""
        nav_items_html += (
            f'<button class="mnav-item {active}" onclick="__mnavGo(\'{sidebar_text}\')">'
            f'<span class="mnav-icon">{svg}</span>'
            f'<span class="mnav-label">{label}</span>'
            f'</button>'
        )

    html = f"""
    <script>
    (function() {{
        if (window.parent.innerWidth > 768) return;

        // Detect dark mode from Streamlit's app background
        function isDark() {{
            var app = window.parent.document.querySelector('[data-testid="stApp"]');
            if (!app) return false;
            var bg = window.parent.getComputedStyle(app).backgroundColor;
            var parts = bg.replace('rgb(','').replace(')','').split(',').map(Number);
            return parts.length === 3 ? (parts[0]+parts[1]+parts[2])/3 < 100 : false;
        }}

        var dark = isDark();
        var navBg      = dark ? '#1a1a2e' : '#ffffff';
        var navBorder  = dark ? '#2e2e3e' : '#e5e5e5';
        var iconColor  = dark ? '#666'    : '#aaa';
        var activeColor = '#e24b4a';

        var existing = window.parent.document.getElementById('__mnav');
        if (existing) existing.remove();

        window.parent.__mnavGo = function(text) {{
            var btns = window.parent.document.querySelectorAll('[data-testid="stSidebar"] button');
            for (var i = 0; i < btns.length; i++) {{
                if (btns[i].textContent.includes(text)) {{ btns[i].click(); return; }}
            }}
        }};

        var nav = window.parent.document.createElement('div');
        nav.id = '__mnav';
        nav.innerHTML = `
            <style>
            #__mnav {{
                position:fixed;bottom:0;left:0;right:0;
                background:${{navBg}};
                border-top:1px solid ${{navBorder}};
                display:flex;height:62px;z-index:999999;
                box-shadow:0 -2px 12px rgba(0,0,0,.12);
            }}
            #__mnav .mnav-item {{
                flex:1;display:flex;flex-direction:column;
                align-items:center;justify-content:center;
                cursor:pointer;border:none;background:none;
                color:${{iconColor}};gap:3px;padding:6px 2px;
                -webkit-tap-highlight-color:transparent;
                transition:color .15s;
            }}
            #__mnav .mnav-item.active {{ color:${{activeColor}}; }}
            #__mnav .mnav-icon svg {{ display:block; }}
            #__mnav .mnav-label {{ font-size:9px;font-family:-apple-system,sans-serif;letter-spacing:.3px; }}
            </style>
            {nav_items_html}
        `;
        window.parent.document.body.appendChild(nav);

        var main = window.parent.document.querySelector('.main .block-container');
        if (main) main.style.paddingBottom = '80px';
    }})();
    </script>
    """
    components.html(html, height=0)


# ─────────────────────────────────────────────
#  ROUTER
# ─────────────────────────────────────────────

pages = {
    "home":            page_home,
    "upload":          page_upload,
    "review":          page_review,
    "results":         page_results,
    "catalog":         page_catalog,
    "wolt":            page_wolt,
    "wolt_export":     page_wolt_export,
    "image_processor": page_image_processor,
    "product_extract": page_product_extract,
    "suppliers":       page_suppliers,
}

pages[st.session_state.page]()
mobile_bottom_nav()