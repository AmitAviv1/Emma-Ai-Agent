"""
wolt_catalog.py — Wolt product catalog loader and fuzzy matcher.

Loads your Wolt catalog Excel file and lets the agent match extracted
invoice products to their Wolt listing by barcode, SKU, or name similarity.

SETUP:
  Add to .env:
    WOLT_CATALOG_PATH=/path/to/your-wolt-catalog.xlsx

  Re-upload a new version of the file any time — just run:
    from wolt_catalog import reload_catalog; reload_catalog()

MATCHING LOGIC (in priority order):
  1. Exact barcode (gtin) match
  2. Exact merchant_sku match
  3. Fuzzy name match (similarity ≥ FUZZY_THRESHOLD)
  4. No match → returns None
"""

import os
import re
import csv
import json
import math
from difflib import SequenceMatcher
from typing import Optional
from dotenv import load_dotenv

load_dotenv()

# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────

WOLT_CATALOG_PATH = os.getenv(
    "WOLT_CATALOG_PATH",
    ""   # Set in .env — e.g. /Users/yourname/Desktop/Projects/AI AGENT/wolt_catalog.xlsx
)

# Minimum similarity score (0-1) to accept a name match
# 0.6 = fairly lenient, 0.75 = strict
FUZZY_THRESHOLD = 0.62

# ── Store-price formula: Wolt price −27%, rounded UP to the shekel ──
STORE_DISCOUNT_1 = 0.27

def compute_store_price(wolt_price: float) -> int:
    if not wolt_price or wolt_price <= 0:
        return 0
    return math.ceil(wolt_price * (1 - STORE_DISCOUNT_1))

# ── Sidecar data files ──
_HERE          = os.path.dirname(os.path.abspath(__file__))
_PROJECT       = os.path.dirname(_HERE)
CATEGORIES_CSV = os.path.join(_PROJECT, "Wolt_Catalog", "categories.csv")
PREV_XLSX      = os.path.join(_PROJECT, "Wolt_Catalog", "wolt_catalog.xlsx")  # previous export, for category recovery
BUY_PRICES     = os.path.join(_PROJECT, "buy_prices.json")         # product_id -> buy price
UNCATEGORIZED  = "ללא קטגוריה"


def _load_json(path: str) -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _load_category_lookup():
    """Read categories.csv → (embedded_id → name, name → parent_name).
    Categories come straight from the file's `category_id` column; no guessing."""
    id_name, child_parent, rows = {}, {}, []
    try:
        with open(CATEGORIES_CSV, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
    except Exception:
        return id_name, child_parent
    for r in rows:
        id_name[r["id"].strip()] = (r.get("name") or "").strip()
    for r in rows:
        subs = (r.get("subcategories") or "").replace("\n", "").split(",")
        for sub_id in subs:
            sub_id = sub_id.strip()
            if sub_id and sub_id in id_name:
                child_parent[id_name[sub_id]] = (r.get("name") or "").strip()
    return id_name, child_parent


def _resolve_category(category_id, id_name, child_parent):
    """Turn a raw `category_id` cell ("name::id" or "NO_CATEGORY::…") into
    (category, subcategory) using categories.csv. Unknown → Uncategorized."""
    if not category_id:
        return UNCATEGORIZED, ""
    embedded_id = str(category_id).split("::")[-1].strip()
    name = id_name.get(embedded_id)
    if not name:
        return UNCATEGORIZED, ""
    parent = child_parent.get(name)
    if parent:
        return parent, name          # leaf → top-level parent + subcategory
    return name, ""


def _load_prev_categories(id_name, child_parent):
    """Recover categories from the PREVIOUS export (wolt_catalog.xlsx) for
    products the current file left uncategorized. Returns dicts keyed by
    product id / gtin / merchant_sku → (category, subcategory).
    Categories a merchant has since removed (not in categories.csv) are kept
    under their own name from the old file — this is real data, not a guess."""
    from openpyxl import load_workbook
    by_id, by_gtin, by_sku = {}, {}, {}
    if not os.path.exists(PREV_XLSX):
        return by_id, by_gtin, by_sku
    wb = load_workbook(PREV_XLSX, read_only=True, data_only=True)
    ws = wb["offers"]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    ci, ii = idx.get("category_id"), idx.get("id")
    gi, si = idx.get("gtin"), idx.get("merchant_sku")
    if ci is None:
        wb.close()
        return by_id, by_gtin, by_sku
    for row in rows:
        if not row or not row[ci]:
            continue
        raw = str(row[ci])
        if raw.startswith("NO_CATEGORY"):
            continue
        cat, sub = _resolve_category(raw, id_name, child_parent)
        if cat == UNCATEGORIZED:
            # a category no longer listed in categories.csv → use its old name
            namepart = raw.split("::")[0].replace("_", " ").strip()
            if not namepart:
                continue
            cat, sub = namepart, ""
        pair = (cat, sub)
        if ii is not None and row[ii]:
            by_id[row[ii]] = pair
        if gi is not None and row[gi]:
            by_gtin[str(row[gi]).strip()] = pair
        if si is not None and row[si]:
            by_sku[str(row[si]).strip()] = pair
    wb.close()
    return by_id, by_gtin, by_sku


def _load_buy_prices() -> dict:
    return _load_json(BUY_PRICES)


def _save_buy_prices(data: dict):
    with open(BUY_PRICES, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─────────────────────────────────────────────
#  CATALOG STRUCTURE
# ─────────────────────────────────────────────

def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return True
    return str(v).strip().lower() in ("1", "true", "yes", "כן")


def _to_float(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


class WoltProduct:
    """One product from the Wolt/Emma catalog (built from a header→value dict)."""
    __slots__ = ("wolt_id", "barcode", "merchant_sku", "name", "description",
                 "sell_price", "store_price", "buy_price", "enabled",
                 "category", "subcategory", "image_urls",
                 "weight_in_grams", "volume_in_ml", "number_of_units")

    def __init__(self, rec: dict):
        # `rec` maps header name → cell value (robust to column reordering).
        self.wolt_id      = rec.get("id") or ""
        self.barcode      = str(rec.get("gtin")).strip() if rec.get("gtin") else ""
        self.merchant_sku = str(rec.get("merchant_sku")).strip() if rec.get("merchant_sku") else ""
        self.name         = rec.get("name") or ""
        self.description  = rec.get("description") or ""
        self.sell_price   = _to_float(rec.get("price"))
        self.store_price  = compute_store_price(self.sell_price)
        self.buy_price    = 0.0                      # filled from buy_prices.json by the loader
        self.enabled      = _to_bool(rec.get("enabled"))
        self.category     = ""                       # filled from the file's category_id by the loader
        self.subcategory  = ""
        images            = rec.get("images") or ""
        self.image_urls   = [u.strip() for u in str(images).split(",") if u.strip()]
        self.weight_in_grams = _to_float(rec.get("weight_in_grams"))
        self.volume_in_ml    = _to_float(rec.get("volume_in_ml"))
        self.number_of_units = _to_float(rec.get("number_of_units"))

    @property
    def image_url(self) -> str:
        """First image (backwards-compatible with old single-image callers)."""
        return self.image_urls[0] if self.image_urls else ""

    def thumb(self, w: int = 300) -> str:
        """Sized image URL via the Wolt image proxy (supports ?w=&h=)."""
        u = self.image_url
        if not u:
            return ""
        sep = "&" if "?" in u else "?"
        return f"{u}{sep}w={w}"

    def to_dict(self) -> dict:
        return {
            "wolt_id":        self.wolt_id,
            "wolt_name":      self.name,
            "wolt_price":     self.sell_price,
            "store_price":    self.store_price,
            "buy_price":      self.buy_price,
            "wolt_enabled":   self.enabled,
            "wolt_barcode":   self.barcode,
            "wolt_sku":       self.merchant_sku,
            "wolt_category":  self.category,
            "wolt_subcategory": self.subcategory,
            "wolt_image":     self.image_url,
        }


# ─────────────────────────────────────────────
#  CATALOG LOADER
# ─────────────────────────────────────────────

_catalog: Optional[list] = None           # list of WoltProduct
_barcode_index: dict = {}                  # barcode → WoltProduct
_sku_index: dict = {}                      # merchant_sku → WoltProduct
_name_tokens: list = []                    # list of (normalized_name, WoltProduct)
_id_index: dict = {}                        # wolt_id → WoltProduct
_category_tree: list = []                   # ordered [(parent, [sub, ...]), ...] for filters


def _normalize(text: str) -> str:
    """Lowercase, strip punctuation and common noise words for fuzzy matching."""
    text = text.lower()
    # Remove size/weight suffixes that vary between invoice and Wolt name
    text = re.sub(r'\|.*$', '', text)            # strip "| 250 גרם" suffix
    text = re.sub(r'\d+[\.,]\d*\s*(ק"ג|קג|גרם|מ"ל|ליטר|יח)', '', text)
    text = re.sub(r'[^\u0590-\u05FFa-zA-Z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _load_from_file(path: str) -> list:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["offers"]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]

    id_name, child_parent = _load_category_lookup()
    fb_id, fb_gtin, fb_sku = _load_prev_categories(id_name, child_parent)
    buy_map = _load_buy_prices()

    products = []
    for row in rows:
        if not row or not any(row):
            continue
        try:
            rec = dict(zip(header, row))
            p = WoltProduct(rec)
            # Category comes directly from the file's category_id column.
            p.category, p.subcategory = _resolve_category(
                rec.get("category_id"), id_name, child_parent)
            # Fill gaps from the previous export (id → gtin → sku).
            if p.category == UNCATEGORIZED:
                pair = (fb_id.get(p.wolt_id)
                        or (fb_gtin.get(p.barcode) if p.barcode else None)
                        or (fb_sku.get(p.merchant_sku) if p.merchant_sku else None))
                if pair:
                    p.category, p.subcategory = pair
            bp = buy_map.get(p.wolt_id)
            if bp is not None:
                p.buy_price = _to_float(bp)
            products.append(p)
        except Exception:
            pass
    wb.close()
    return products


def _load_category_tree() -> list:
    """Ordered [(parent_name, [subcategory_name, ...]), ...] from categories.csv,
    used to render the category filter grouped parent → sub."""
    tree = []
    id_name = {}
    try:
        with open(CATEGORIES_CSV, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
    except Exception:
        return tree
    for r in rows:
        id_name[r["id"].strip()] = (r.get("name") or "").strip()
    child_ids = set()
    for r in rows:
        subs = [s.strip() for s in (r.get("subcategories") or "").replace("\n", "").split(",") if s.strip()]
        if subs:
            children = [id_name[s] for s in subs if s in id_name]
            tree.append(((r.get("name") or "").strip(), children))
            child_ids.update(subs)
    # top-level categories that are not a parent-with-children and not a child
    for r in rows:
        cid = r["id"].strip()
        name = (r.get("name") or "").strip()
        is_parent = any(name == p for p, _ in tree)
        if cid not in child_ids and not is_parent:
            tree.append((name, []))
    return tree


def _build_indexes(products: list):
    global _barcode_index, _sku_index, _name_tokens, _id_index, _category_tree
    _barcode_index = {}
    _sku_index     = {}
    _name_tokens   = []
    _id_index      = {}

    for p in products:
        if p.barcode:
            _barcode_index[p.barcode] = p
        if p.merchant_sku:
            _sku_index[p.merchant_sku] = p
        if p.wolt_id:
            _id_index[p.wolt_id] = p
        _name_tokens.append((_normalize(p.name), p))

    _category_tree = _load_category_tree()

    print(f"📦 [WOLT] Catalog loaded: {len(products)} products "
          f"({len(_barcode_index)} with barcode, "
          f"{len(_sku_index)} with SKU)")


def load_catalog(path: str = "") -> bool:
    """
    Load (or reload) the Wolt catalog from an Excel file.
    Returns True on success, False if file not found.
    """
    global _catalog
    path = path or WOLT_CATALOG_PATH
    if not path or not os.path.exists(path):
        print(f"⚠️  [WOLT] Catalog file not found: '{path}'. "
              f"Set WOLT_CATALOG_PATH in .env")
        return False

    print(f"📂 [WOLT] Loading catalog from {os.path.basename(path)}…")
    _catalog = _load_from_file(path)
    _build_indexes(_catalog)
    return True


def reload_catalog():
    """Force reload — call this after uploading a new Wolt export."""
    return load_catalog()


def is_loaded() -> bool:
    return _catalog is not None and len(_catalog) > 0


def _ensure_loaded():
    if not is_loaded():
        load_catalog()


# ─────────────────────────────────────────────
#  PUBLIC ACCESSORS (for the Streamlit shop catalog)
# ─────────────────────────────────────────────

def all_products() -> list:
    """Return the loaded product list (empty list if not loaded)."""
    _ensure_loaded()
    return _catalog or []


def category_tree() -> list:
    """Ordered [(parent_name, [subcategory, ...]), ...] for the category filter."""
    _ensure_loaded()
    return _category_tree


def categories() -> list:
    """Flat, de-duplicated list of category names actually present in the catalog,
    ordered by the category tree (parents first), with any extras appended."""
    _ensure_loaded()
    present = {p.category for p in (_catalog or []) if p.category}
    ordered = []
    for parent, subs in _category_tree:
        if parent in present and parent not in ordered:
            ordered.append(parent)
    for c in sorted(present):
        if c not in ordered:
            ordered.append(c)
    return ordered


def set_buy_price(product_id: str, value) -> bool:
    """Persist a buy price for a product and update the in-memory object."""
    if not product_id:
        return False
    data = _load_buy_prices()
    try:
        val = float(value)
    except (TypeError, ValueError):
        return False
    if val > 0:
        data[product_id] = val
    else:
        data.pop(product_id, None)
        val = 0.0
    _save_buy_prices(data)
    p = _id_index.get(product_id)
    if p is not None:
        p.buy_price = val
    return True


# ─────────────────────────────────────────────
#  MATCHING
# ─────────────────────────────────────────────

def _fuzzy_score(a: str, b: str) -> float:
    """SequenceMatcher similarity ratio between two normalized strings."""
    return SequenceMatcher(None, a, b).ratio()


def match_product(
    barcode: str = "",
    sku: str     = "",
    name: str    = "",
) -> Optional[dict]:
    """
    Find the best Wolt catalog match for an invoice product.

    Priority:
      1. Exact barcode match
      2. Exact SKU match
      3. Best fuzzy name match (if score ≥ FUZZY_THRESHOLD)

    Returns a dict with wolt_* fields, or None if no match found.
    Also includes match_method and match_score for transparency.
    """
    _ensure_loaded()

    if not is_loaded():
        return None

    # 1. Exact barcode
    if barcode and barcode in _barcode_index:
        p = _barcode_index[barcode]
        return {**p.to_dict(), "match_method": "barcode", "match_score": 1.0}

    # 2. Exact SKU (merchant_sku in Wolt = often barcode or internal code)
    if sku and sku in _sku_index:
        p = _sku_index[sku]
        return {**p.to_dict(), "match_method": "sku", "match_score": 1.0}

    # Also try barcode against SKU index (Wolt sometimes stores barcode in merchant_sku)
    if barcode and barcode in _sku_index:
        p = _sku_index[barcode]
        return {**p.to_dict(), "match_method": "barcode_as_sku", "match_score": 1.0}

    # 3. Fuzzy name match
    if name:
        norm_name = _normalize(name)
        best_score = 0.0
        best_match = None

        for norm_wolt, p in _name_tokens:
            score = _fuzzy_score(norm_name, norm_wolt)
            if score > best_score:
                best_score = score
                best_match = p

        if best_score >= FUZZY_THRESHOLD and best_match:
            return {
                **best_match.to_dict(),
                "match_method": "name_fuzzy",
                "match_score": round(best_score, 3),
            }

    return None


def enrich_products(products: list) -> list:
    """
    Add Wolt catalog data to a list of extracted invoice products.
    Each product dict gets wolt_* fields and a margin calculation.

    Returns the enriched list (modifies in place).
    """
    _ensure_loaded()
    matched = 0

    for p in products:
        wolt = match_product(
            barcode=p.get("barcode", ""),
            sku=p.get("sku", ""),
            name=p.get("name", ""),
        )

        if wolt:
            matched += 1
            p.update(wolt)

            # Calculate margin if both prices available
            sell  = wolt.get("wolt_price", 0)
            cost  = p.get("cost", 0)
            qty   = p.get("quantity", 1) or 1
            unit_cost = cost / qty if qty else cost

            if sell > 0 and unit_cost > 0:
                margin_pct = round((sell - unit_cost) / sell * 100, 1)
                p["margin_pct"]  = margin_pct
                p["unit_cost"]   = round(unit_cost, 2)
                p["wolt_price"]  = sell
        else:
            p["match_method"] = "none"
            p["match_score"]  = 0.0

    print(f"🔗 [WOLT] Matched {matched}/{len(products)} products to Wolt catalog")
    return products


# ─────────────────────────────────────────────
#  CATALOG UPLOAD HELPER (for Streamlit UI)
# ─────────────────────────────────────────────

def load_catalog_from_bytes(file_bytes: bytes, filename: str) -> bool:
    """
    Load catalog directly from uploaded file bytes (for Streamlit uploader).
    Saves to a temp path and loads from there.
    """
    import tempfile
    suffix = os.path.splitext(filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    result = load_catalog(tmp_path)
    os.unlink(tmp_path)
    return result